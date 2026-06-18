"""
Generador de planillas Excel para todas las actividades.
Uso: generate(params, output_dir) → path del archivo generado

params keys: activity, site, pax, date, guide, client, hotel,
             hora, gastronomy_type, dietary_restrictions
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Estilos ────────────────────────────────────────────────────────────────────
_fl = Font(name="Arial", bold=True, size=10)
_fn = Font(name="Arial", size=10)
_fi = Font(name="Arial", size=10, color="0000CC")
_fa = Font(name="Arial", size=10, color="1A6B1A")
_fs = Font(name="Arial", size=9, color="777777", italic=True)

_sub  = PatternFill("solid", start_color="D6E4F0")
_inp  = PatternFill("solid", start_color="FFF9C4")
_auto = PatternFill("solid", start_color="E8F5E9")
_wht  = PatternFill("solid", start_color="FFFFFF")
_opt  = PatternFill("solid", start_color="F5F5F5")

_al = Alignment(horizontal="left",   vertical="center")
_ac = Alignment(horizontal="center", vertical="center")
_th = Side(style="thin", color="AAAAAA")
_bd = Border(left=_th, right=_th, top=_th, bottom=_th)

PAX = "$B$6"

# ── Helpers ────────────────────────────────────────────────────────────────────
def _sec(ws, row, text):
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value, c.font, c.alignment = text, Font(name="Arial", bold=True, size=10, color="FFFFFF"), _al
    c.fill = PatternFill("solid", start_color="2E4057")
    ws.row_dimensions[row].height = 18

def _hdr(ws, row):
    for col, txt in [("A","Item"),("B","Cantidad")]:
        c = ws[f"{col}{row}"]
        c.value, c.font, c.alignment, c.border = txt, Font(name="Arial", bold=True, size=10, color="FFFFFF"), _ac, _bd
        c.fill = PatternFill("solid", start_color="4A7FB5")
    ws.row_dimensions[row].height = 18

def _lv(ws, row, label, value="", mode="auto"):
    ws[f"A{row}"] = label
    ws[f"A{row}"].font, ws[f"A{row}"].fill = _fl, _sub
    ws[f"A{row}"].alignment, ws[f"A{row}"].border = _al, _bd
    ws[f"B{row}"] = value
    ws[f"B{row}"].font  = _fa if mode == "auto" else _fi
    ws[f"B{row}"].fill  = _auto if mode == "auto" else _inp
    ws[f"B{row}"].alignment, ws[f"B{row}"].border = _al, _bd
    ws.row_dimensions[row].height = 16

def _it(ws, row, label, val, mode="auto"):
    """
    mode: auto=formula/fijo, input=celda amarilla editable, optional=gris 'consultar con guía'
    """
    ws[f"A{row}"] = label
    ws[f"A{row}"].font, ws[f"A{row}"].alignment, ws[f"A{row}"].border = _fn, _al, _bd

    ws[f"B{row}"] = val
    ws[f"B{row}"].alignment, ws[f"B{row}"].border = _ac, _bd

    if mode == "input":
        ws[f"B{row}"].font, ws[f"B{row}"].fill = _fi, _inp
    elif mode == "optional":
        ws[f"B{row}"].font, ws[f"B{row}"].fill = _fs, _opt
    else:
        ws[f"B{row}"].font, ws[f"B{row}"].fill = _fn, _wht
    ws.row_dimensions[row].height = 16

def _gap(ws, row, h=6):
    ws.row_dimensions[row].height = h

def _drinks(ws, start_row):
    _sec(ws, start_row, "  BEBIDAS")
    _hdr(ws, start_row + 1)
    items = [
        ("Agua 600ml",       f"={PAX}"),
        ("Agua 1.5L",        f"=INT({PAX}/2)"),
        ("Cerveza",          f"={PAX}+1"),
        ("Vinos (botellas)", f"=CEILING({PAX}/2,1)"),
    ]
    for i, (lbl, val) in enumerate(items):
        _it(ws, start_row + 2 + i, lbl, val)
    return start_row + 2 + len(items)

def _cabecera(ws, params):
    fecha = params.get('date')
    fecha_str = fecha.strftime('%d/%m/%Y') if isinstance(fecha, date) else (fecha or '')
    _sec(ws, 1, "  DATOS DE LA SALIDA")
    _lv(ws, 2,  "SALIDA:",            params.get('activity') or '', mode="auto")
    _lv(ws, 3,  "SITIO:",             params.get('site') or '',     mode="input")
    _lv(ws, 4,  "FECHA:",             fecha_str,                    mode="auto")
    _lv(ws, 5,  "HORA DE RETIRO:",    params.get('hora') or '',     mode="input")
    _lv(ws, 6,  "CANT. PAX:",         params.get('pax') or 0,      mode="auto")
    _lv(ws, 7,  "NOMBRE / CONTACTO:", params.get('client') or '',   mode="input")
    _lv(ws, 8,  "CLIENTE / AGENCIA:", '',                           mode="input")
    _lv(ws, 9,  "HOTEL:",             params.get('hotel') or '',    mode="input")
    _lv(ws, 10, "GUIA:",              params.get('guide') or '',    mode="input")
    _lv(ws, 11, "CONDUCTOR:",         '',                           mode="input")
    _lv(ws, 12, "RESTRICCIONES:",     params.get('dietary_restrictions') or 'Ninguna', mode="auto")
    ws["A6"].font = Font(name="Arial", bold=True, size=10, color="CC0000")
    ws["B6"].font = Font(name="Arial", bold=True, size=11, color="1A6B1A")
    ws["B6"].fill = _auto
    _gap(ws, 13)

# ── Templates de secciones ─────────────────────────────────────────────────────
# Cada item: (label, value, mode)
# mode: "auto" | "input" | "optional"
# "optional" → value se ignora, muestra "consultar con guía"
OPT = "consultar con guía"

def _get_almuerzo_campestre_rows(mesa4_row, mesa2_row):
    """Items del almuerzo campestre. mesa4_row y mesa2_row son las filas de las mesas para referencia."""
    m4 = f"B{mesa4_row}"
    m2 = f"B{mesa2_row}"
    return [
        (f"Mesa roja (de 4 personas)",             f"=INT(({PAX}+1)/4)",                              "auto"),
        (f"Mesa (de 2 personas)",                  f"=IF(AND(MOD({PAX},4)>=1,MOD({PAX},4)<=2),1,0)", "auto"),
        ("Sillitas de camping azules",             f"={PAX}",                                         "auto"),
        ("Mantitas norteñas",                      f"={PAX}",                                         "auto"),
        ("Manta cuadros turquesa",                 "1",                                               "auto"),
        ("Copas grandes (vino)",                   f"={PAX}",                                         "auto"),
        ("Copas chicas (agua)",                    f"={PAX}",                                         "auto"),
        ("Tenedores",                              f"={PAX}",                                         "auto"),
        ("Cucharitas (postre / cafe)",             f"={PAX}",                                         "auto"),
        ("Cuchillos",                              f"={PAX}",                                         "auto"),
        ("Cuchara ensalada",                       "1 juego",                                         "auto"),
        ("Destapador",                             "1",                                               "auto"),
        ("Cuencos ceramicos (sopa/dips)",          f"=IF({PAX}<=4,2,IF({PAX}<=7,3,4))",              "auto"),
        ("Cucharita madera",                       f"=IF({PAX}<=4,2,IF({PAX}<=7,3,4))",              "auto"),
        ("Servilletas / pax",                      "varios",                                          "auto"),
        ("Servilletas rollo cocina",               "varios",                                          "auto"),
        ("Servilleteros",                          "1",                                               "auto"),
        ("Paneras tela",                           "1",                                               "auto"),
        ("Individuales tela (solo Barco)",         "/",                                               "auto"),
        ("Platos principal",                       f"={PAX}",                                         "auto"),
        ("Platos ceramicos kawen (postre)",        f"=IF({PAX}<=4,1,IF({PAX}<=8,2,3))",              "auto"),
        ("Tablas de madera",                       f"=CEILING({PAX}/2,1)",                            "auto"),
        ("Ensaladera madera",                      "1",                                               "auto"),
        ("Broches mantel",                         f"={m4}+{m2}",                                    "auto"),
        ("Manteles",                               f"={m4}+{m2}",                                    "auto"),
        ("Caminos",                                f"={m4}+{m2}",                                    "auto"),
        ("Alcohol gel",                            "1",                                               "auto"),
        ("Bolsas de basura",                       "varios",                                          "auto"),
    ]

TEMPLATES = {

    'FD Kayak': {
        'gastronomy_title': 'GASTRONOMIA — Almuerzo Campestre FD',
        'equipment': [
            ("Guantes (pares)",                    f"={PAX}",   "auto"),
            ("Mitones para los remos (sin talle)", "2",         "auto"),
            ("Botitas neopreno (talles)",          "",          "input"),
            ("Chaleco / PFD (talles)",            "",          "input"),
            ("Chaqueta impermeable (talles)",     "",          "input"),
            ("Copit / cubre (talle unico)",       f"={PAX}",   "auto"),
            ("Culotes / asientos kayak",           f"={PAX}+1", "auto"),
            ("Bomba achique",                      "1",         "auto"),
            ("Bolsas secas",                       "1",         "auto"),
            ("Remos",                              f"={PAX}",   "auto"),
            ("Remos / palas desmontables negras",  f"={PAX}+1", "auto"),
            ("Radio / Handys",                     "2",         "auto"),
        ],
        'gastronomy': None,   # se genera dinámicamente (almuerzo campestre)
        'gastronomy_type': 'AC',
    },

    'HD Kayak': {
        'gastronomy_title': 'GASTRONOMIA — Snacks HD',
        'equipment': [
            ("Guantes (pares)",                    f"={PAX}",   "auto"),
            ("Mitones para los remos (sin talle)", "2",         "auto"),
            ("Botitas neopreno (talles)",          "",          "input"),
            ("Chaleco / PFD (talles)",            "",          "input"),
            ("Chaqueta impermeable (talles)",     "",          "input"),
            ("Copit / cubre (talle unico)",       f"={PAX}",   "auto"),
            ("Culotes / asientos kayak",           f"={PAX}",   "auto"),
            ("Bomba achique",                      "1",         "auto"),
            ("Bolsas secas",                       "3",         "auto"),
            ("Remos",                              f"={PAX}",   "auto"),
            ("Remos / palas desmontables negras",  "1",         "auto"),
            ("Radio / Handys",                     "2",         "auto"),
        ],
        'gastronomy': [
            ("Tazas te",                           f"={PAX}",   "auto"),
            ("Cucharitas te",                      f"={PAX}",   "auto"),
            ("Caja de te madera",                  "1",         "auto"),
            ("Termo agua",                         "1",         "auto"),
            ("Mesa pequeña",                       "1",         "auto"),
            ("Mantel verde",                       "1",         "auto"),
            ("Mantitas norteñas",                  f"={PAX}",   "auto"),
            ("Sillitas camping",                   f"={PAX}",   "auto"),
            ("Bolsas de residuos",                 "varios",    "auto"),
            ("Servilletas",                        "varios",    "auto"),
            ("Alcohol gel",                        "1",         "auto"),
            ("Pinza comida",                       "1",         "auto"),
            ("Cuchillo",                           "1",         "auto"),
            ("Tabla de madera + ceramica kawen",   "1 + 1",     "auto"),
            ("— COMIDA —",                         "",          "auto"),
            ("Agua",                               f"={PAX}",   "auto"),
            ("Brownie",                            f"={PAX}",   "auto"),
            ("Budin",                              "varios",    "auto"),
            ("Cookies",                            "varios",    "auto"),
            ("Sandwich de miga",                   "varios",    "auto"),
        ],
        'gastronomy_type': 'snacks',
        'no_drinks': True,   # bebidas integradas en snacks section
    },

    'FD MTB': {
        'gastronomy_title': 'GASTRONOMIA — Box Lunch',
        'equipment': [
            ("Guantes",         f"={PAX}", "auto"),
            ("Ponchos lluvia",  f"={PAX}", "auto"),
            ("Radios",          "2",       "auto"),
        ],
        'gastronomy': [
            ("BLTREK",          f"={PAX}", "auto"),
            ("Personal (guia + conductor — ajustar si hay backup)", "2", "input"),
        ],
        'gastronomy_type': 'BL',
    },

    'FD Trekking': {
        'gastronomy_title': 'GASTRONOMIA — Box Lunch',
        'equipment': [
            ("Bastones",  f"={PAX}",  "auto"),
            ("Radios",    OPT,        "optional"),
        ],
        'gastronomy': [
            ("BLTREK",   f"={PAX}", "auto"),
            ("Personal (guia + conductor — ajustar si hay backup)", "2", "input"),
        ],
        'gastronomy_type': 'BL',
    },

    'Montañas y Glaciares del Tronador': {
        'gastronomy_title': 'GASTRONOMIA — Box Lunch Gourmet',
        'equipment': [
            ("Bastones",            OPT,       "optional"),
            ("Radios",              OPT,       "optional"),
            ("Vasos",               f"={PAX}", "auto"),
            ("Sacacorchos",         "1",       "auto"),
            ("Mesa",                OPT,       "optional"),
            ("Mantel",              OPT,       "optional"),
            ("Sillitas de camping", OPT,       "optional"),
        ],
        'gastronomy': [
            ("BLgourmet",  f"={PAX}", "auto"),
            ("Personal (guia + conductor — ajustar si hay backup)", "2", "input"),
        ],
        'gastronomy_type': 'BL',
    },

    'Rios y Lagos del Limay': {
        'gastronomy_title': 'GASTRONOMIA — Box Lunch Gourmet',
        'equipment': [
            ("Bastones",  OPT,       "optional"),
            ("Mantitas",  f"={PAX}", "auto"),
            ("Radios",    "2",       "auto"),
        ],
        'gastronomy': [
            ("BLgourmet",  f"={PAX}", "auto"),
            ("Personal (guia + conductor — ajustar si hay backup)", "2", "input"),
        ],
        'gastronomy_type': 'BL',
    },

    'Navegacion Regular': {
        'gastronomy_title': 'GASTRONOMIA',
        'equipment': [
            ("Bastones",          OPT, "optional"),
            ("Radios",            OPT, "optional"),
            ("Ponchos de lluvia", OPT, "optional"),
        ],
        'gastronomy': [
            ("Box Lunch",         f"={PAX}", "auto"),
            ("Personal (guia + conductor — ajustar si hay backup)", "2", "input"),
            ("Bebidas adicionales", "",      "input"),
        ],
        'gastronomy_type': 'BL',
    },

    'Navegacion AC': {
        'gastronomy_title': 'GASTRONOMIA — Almuerzo Campestre (Barco)',
        'equipment': [
            ("Bastones",          OPT, "optional"),
            ("Radios",            OPT, "optional"),
            ("Ponchos de lluvia", OPT, "optional"),
        ],
        'gastronomy': [
            # Navegación AC: sin mesas (es en el barco), individuales = 1/pax
            ("Copas grandes (vino)",               f"={PAX}",                           "auto"),
            ("Copas chicas (agua)",                f"={PAX}",                           "auto"),
            ("Tenedores",                          f"={PAX}",                           "auto"),
            ("Cucharitas (postre / cafe)",         f"={PAX}",                           "auto"),
            ("Cuchillos",                          f"={PAX}",                           "auto"),
            ("Cuchara ensalada",                   "1 juego",                           "auto"),
            ("Destapador",                         "1",                                 "auto"),
            ("Cuencos ceramicos (dips)",           f"={PAX}",                           "auto"),
            ("Cucharita madera / dips",            f"={PAX}",                           "auto"),
            ("Servilletas / pax",                  "varios",                            "auto"),
            ("Servilletas rollo cocina",           "varios",                            "auto"),
            ("Servilleteros",                      "1",                                 "auto"),
            ("Paneras tela",                       "1",                                 "auto"),
            ("Individuales tela",                  f"={PAX}",                           "auto"),
            ("Platos principal",                   f"={PAX}",                           "auto"),
            ("Platos ceramicos kawen (postre)",    f"=CEILING({PAX}/2,1)",              "auto"),
            ("Tablas de madera",                   f"=CEILING({PAX}/2,1)",              "auto"),
            ("Ensaladera madera grande",           "1",                                 "auto"),
            ("Broches mantel",                     "4",                                 "input"),
            ("Manteles",                           "1",                                 "input"),
            ("Caminos",                            "1",                                 "input"),
            ("Alcohol gel",                        "1",                                 "auto"),
            ("Bolsas de basura",                   "varios",                            "auto"),
        ],
        'gastronomy_type': 'AC',
    },

    'Valle Asado': {
        'gastronomy_title': 'CATERING — Asado Valle Encantado',
        'equipment': [
            ("Bastones / Radios", OPT, "optional"),
            ("Heladera comida",   "1", "auto"),
            ("Heladera bebidas",  "1", "auto"),
        ],
        'gastronomy': [
            ("Empanadas (humita)",              f"={PAX}",                "auto"),
            ("Colita de cuadril",               f"=CEILING({PAX}/2,1)",  "auto"),
            ("Chori de cerdo",                  f"={PAX}",                "auto"),
            ("Morcilla",                        f"={PAX}",                "auto"),
            ("Cerdo",                           f"=CEILING({PAX}/2,1)",  "auto"),
            ("Pan",                             "1 kg",                   "input"),
            ("Ensaladas",                       "varias",                 "input"),
            ("Verduras para parrilla",          "varias",                 "input"),
            ("Dulce de leche",                  "1",                      "auto"),
            ("Postre / Brownie",                f"={PAX}",                "auto"),
            ("Crema",                           "1",                      "auto"),
        ],
        'gastronomy_type': 'asado',
    },
}

# Alias para parsear variantes del nombre
ACTIVITY_ALIASES = {
    'fd kayak y trekking': 'FD Kayak',
    'fd kayak':            'FD Kayak',
    'kayak mascardi':      'FD Kayak',
    'hd kayak':            'HD Kayak',
    'kayak snacks':        'HD Kayak',
    'mtb':                 'FD MTB',
    'fd mtb':              'FD MTB',
    'trekking':            'FD Trekking',
    'fd trekking':         'FD Trekking',
    'tronador':            'Montañas y Glaciares del Tronador',
    'montañas y glaciares':'Montañas y Glaciares del Tronador',
    'limay':               'Rios y Lagos del Limay',
    'rios y lagos':        'Rios y Lagos del Limay',
    'navegacion regular':  'Navegacion Regular',
    'navegacion':          'Navegacion Regular',
    'navegacion ac':       'Navegacion AC',
    'navegacion pvt':      'Navegacion AC',
    'barco':               'Navegacion Regular',
    'valle asado':         'Valle Asado',
    'asado':               'Valle Asado',
}


def resolve_activity(raw: str) -> str:
    """Convierte el nombre parseado al nombre canónico del template."""
    key = raw.lower().strip()
    return ACTIVITY_ALIASES.get(key, raw)


def generate(params: dict, output_dir: str = ".") -> str:
    activity_raw = params.get('activity') or 'FD Kayak'
    activity     = resolve_activity(activity_raw)
    tmpl         = TEMPLATES.get(activity, TEMPLATES['FD Kayak'])

    pax   = params.get('pax') or 0
    fecha = params.get('date')
    fecha_str = fecha.strftime('%d/%m/%Y') if isinstance(fecha, date) else (fecha or '')

    wb = Workbook()
    ws = wb.active
    ws.title = activity[:28]   # Excel limit on sheet name
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 26

    # CABECERA
    _cabecera(ws, {**params, 'activity': activity})

    # EQUIPAMIENTO
    equip = tmpl.get('equipment', [])
    row = 14
    _sec(ws, row, "  EQUIPAMIENTO");   row += 1
    _hdr(ws, row);                     row += 1
    for lbl, val, mode in equip:
        display = OPT if mode == "optional" else val
        _it(ws, row, lbl, display, mode=mode)
        row += 1
    _gap(ws, row); row += 1

    # GASTRONOMIA
    gast_title = tmpl.get('gastronomy_title', 'GASTRONOMIA')
    _sec(ws, row, f"  {gast_title}"); row += 1
    _hdr(ws, row);                    row += 1

    gastronomy = tmpl.get('gastronomy')
    if gastronomy is None:
        # Almuerzo campestre kayak FD (con referencia a filas de mesas)
        mesa4_row = row  # B de esta fila = mesas de 4
        mesa2_row = row + 1
        rows_data = _get_almuerzo_campestre_rows(mesa4_row, mesa2_row)
        for lbl, val, mode in rows_data:
            _it(ws, row, lbl, val, mode=mode)
            row += 1
    else:
        for lbl, val, mode in gastronomy:
            display = OPT if mode == "optional" else val
            _it(ws, row, lbl, display, mode=mode)
            row += 1

    _gap(ws, row); row += 1

    # BEBIDAS (salvo HD Kayak que las integra en snacks)
    if not tmpl.get('no_drinks'):
        _drinks(ws, row)

    # Guardar
    act_slug  = activity.replace(' ', '_').replace('/', '-')
    date_slug = fecha.strftime('%Y%m%d') if isinstance(fecha, date) else 'sin_fecha'
    filename  = f"Planilla_{act_slug}_{date_slug}_{pax}pax.xlsx"
    filepath  = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
